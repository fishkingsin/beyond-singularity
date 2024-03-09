from openpyxl import load_workbook

# Path to your Excel file
xlsx_file_path = 'Beyond Singularity (Responses).xlsx'  # Update this to the path of your Excel file

# Load the workbook and select a worksheet
wb = load_workbook(filename=xlsx_file_path)
sheet = wb.active  # Assumes you're working with the first sheet; update as necessary
headers = [cell.value for cell in sheet[1]]  # Headers are typically in the first row
first_few_rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_row=9)]
def convert_row_to_html(row, row_number):
    # Define the HTML template, using row values like row[1] for name, row[0] for timestamp, etc.
    name = row[1]  # Assuming second column is 'Name'
    
    
    html_template = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta http-equiv="X-UA-Compatible" content="IE=edge">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Submission Detail - {name}</title>
    </head>
    <body>
        <div>
    """
    
    # Iterate over the cells in the row starting from the 3rd column
    for cell in row[2:]:
        if cell:
            link = cell
            file_id = link.split('id=')[-1]
            if file_id and 'drive.google.com' in link:
                html_template += f"""
                    <div>

                        <img src="https://drive.google.com/uc?export=view&id={file_id}" " style="width:100%; max-width:1000px;">
                    </div>
                """
            else:
                html_template += f"<p>{cell}</p>"
    
    html_template += """
        </div>
    </body>
    </html>
    """
    return html_template

# Iterate over each row in the worksheet and generate an HTML file, skipping the header row
for index, row in enumerate(first_few_rows, start=1):  # Adjust min_row as necessary
    html_content = convert_row_to_html(row, index)
    file_name = f"submission_{index}.html"  # Generates a unique file name for each submission
    with open(file_name, 'w', encoding='utf-8-sig') as file:
        file.write(html_content)
    print(f"Generated {file_name}")
