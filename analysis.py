from openpyxl import load_workbook

# Load the workbook
file_path = 'Beyond Singularity (Responses).xlsx'
wb = load_workbook(filename=file_path)

# Assuming we are working with the first sheet
sheet = wb.active

# Let's get an overview by reading the headers and the first few rows of data
headers = [cell.value for cell in sheet[1]]  # Headers are typically in the first row
first_few_rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_row=9)]

headers, first_few_rows
# print(headers)
print(first_few_rows[0][2])